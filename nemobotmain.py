import asyncio
import discord
from discord.ui import Button, View
from discord.ext import commands
import discord.enums
from discord.utils import get
from discord.http import Route
from discord import Intents 
import openpyxl
from openpyxl.styles import numbers
from openpyxl.utils import get_column_letter
from openpyxl import Workbook, load_workbook
from discord import option
import schedule
import datetime
import time
import traceback
from discord.ext import tasks
from datetime import datetime


intents = discord.Intents.all()
intents.guilds = True
client = discord.Client
openxl = openpyxl.load_workbook("nemobot.xlsx")
bot = commands.Bot(command_prefix='!',intents=intents)
dt = datetime.now()
global preset
preset = 0

def get_username(idctx, guild):
    if idctx != 'None':
        id_ctx = str(idctx)
        print(id_ctx)
        user = guild.get_member(int(id_ctx))
        if user is not None:
        # found the user
            print(user.display_name)
            return user.display_name
        else:
        # Not found the user
            print("no user (id)")
            return id_ctx
    else:
         return idctx

def sheetname(servername):
    if servername in openxl.sheetnames:
        if openxl.active.title != servername:
            openxl.move_sheet(servername,0)
        openxl.save('nemobot.xlsx')
        return servername
                        
    elif servername not in openxl.sheetnames:
        openxl.create_sheet('{}'.format(servername),0)
        openxl.save('nemobot.xlsx')
        return servername


def checkRow(ctx):
        for row in range(1, openxl[sheetname(ctx)].max_row + 1):
            if openxl[sheetname(ctx)].cell(row, 1).value is None or '':
                break
        return row +1
        
def checkCol(ctx):
        for column in range(1, openxl[sheetname(ctx)].max_column+1):
            if openxl[sheetname(ctx)].cell(1, column).value is None:
                break
        return column +1

def sheetsort(ctx):
    sheet = openxl[sheetname(ctx)]
    
    _row = checkRow(ctx)
    _col = checkCol(ctx)
    # 데이터 불러오기
    data = []
    for row_num in range(1, _row):
        row_data = []
        for col_num in range(1, _col):
            cell_value = openxl[sheetname(ctx)].cell(row=row_num, column=col_num).value
            if cell_value is not None:
                row_data.append(cell_value)
        data.append(row_data)
    
    print(data)
    # B열을 기준으로 내림차순으로 정렬
    if _row > 1:
        sorted_data = sorted(data, key=lambda x: x[1], reverse=True)
    else:
        sorted_data = data
    
    # 정렬된 데이터를 다시 셀에 입력
    for row_idx, row_data in enumerate(sorted_data, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            sheet.cell(row=row_idx, column=col_idx, value= value)
    openxl.save("nemobot.xlsx")

def editbnr(ctx):
    servername = sheetname(ctx.guild.name)
    editbanner = discord.Embed(title="출석체크!",description="순위", colour=discord.Colour.blue())
    editbanner.set_footer(text="출석하려면 버튼을 눌러주세요")
    sheetsort(servername)#시트소팅    
    
    _row = checkRow(servername)
    for row in range(1, _row):
         value2 = "누적 {} 회 오늘은 {}\n마지막 출석 {}".format (openxl[servername].cell(row, 2).value, openxl[servername].cell(row, 3).value, openxl[servername].cell(row, 4).value)
         try:
             name = get_username(str(openxl[servername].cell(row, 1).value), ctx.guild)
         except:
             name = openxl[servername].cell(row, 1).value
         editbanner.add_field(name=name, value=value2, inline=False)
     
    return editbanner

async def bannersendfnc(channel):
    checkbutton = Button(label="출석확인", style=discord.ButtonStyle.green, custom_id= 'checkbutton')
    
    view = View()
    view.add_item(checkbutton)

    es = discord.Embed(title="출석체크!",description="순위", colour=discord.Colour.blue())
    es.set_footer(text="출석하려면 버튼을 눌러주세요")
    ctx = channel.guild.name
    servername = sheetname(ctx)
    sheetsort(servername)
    _row = checkRow(ctx)  
    for row in range(1, _row):
         if openxl[servername].cell(row, 1).value != None and openxl[servername].cell(row, 4).value != "{}년 {}월 {}일".format (datetime.now().year, datetime.now().month, datetime.now().day):
             openxl[servername].cell(row= row, column=3, value = '출석안함')
         
    for row in range(1, _row):
         value2 = "누적 {} 회 오늘은 {}\n마지막 출석 {}".format (openxl[servername].cell(row, 2).value, openxl[servername].cell(row, 3).value, openxl[servername].cell(row, 4).value)
         name= get_username(str(openxl[servername].cell(row, 1).value), channel.guild)
         es.add_field(name=name, value=value2, inline=False) 
    
    return es, view

async def dailyset():
      for joiningguild in bot.guilds:
          includech = joiningguild.text_channels
          for ch in includech:
              if ch.name == "네모봇출석체크방-beta":
                   target_message=await ch.history(limit=None).flatten()
                   print("daily setting message:")
                   print([target_message])
                   await ch. delete_messages(target_message)
                   value1 ,value2 = await bannersendfnc(ch)
                   global bannersend
                   bannersend=await ch.send(embed=value1, view= value2)
                   break
              else:
               return
          else:
               return 
               
schedule.every().day.at("00:00").do(lambda: asyncio.run(dailyset()))
    
@bot.event
async def on_ready():
    global bannersend
    print('다음으로 로그인합니다: ')
    print(bot.user.name)
    print('connection was succesful')
    await bot.change_presence(status=discord.Status.online, activity=discord.Game("/명령어"))
    for joiningguild in bot.guilds:    
        print("bot is checking channel:")
        for ch in joiningguild.text_channels:
              print(ch.name)
              if ch.name == "네모봇출석체크방-beta":    
                      target_message=await ch.history(limit=None).flatten()
                      print("bot is checking message:")
                      print([target_message])
                      await ch. delete_messages(target_message)
                      value1 ,value2 = await bannersendfnc(ch)
                      await ch.send("봇이 재연결되었습니다", delete_after = 2)
                      global bannersend
                      bannersend=await ch.send(embed=value1, view= value2)
                      break
        else:
            pass   
            
    for joiningguild in bot.guilds:
        roles = await joiningguild.fetch_roles()
        for role in roles:
            if role.name == '네모봇':
                 await joiningguild.get_member(bot.user.id).add_roles(role)
                 break
        else:
             await joiningguild.create_role( name='네모봇', permissions= discord.Permissions(administrator=True), colour=discord.Colour.gold(), hoist=True, mentionable=True, reason="for nemobot attendent channel authority")
             await joiningguild.get_member(bot.user.id).add_roles(discord.utils.get(joiningguild.roles, name='네모봇'))
    else:
         pass
    
    while on_ready:
        res = await bot.wait_for('interaction')
        for item in [res.data]:
           try: 
            if item['custom_id'] == 'checkbutton':
                # 'checkbutton' 뒤의 문자열 추출
                inputres = res
                print("\n{}서버에서 출첵버튼 눌림".format(inputres.guild.name))
                print(inputres.user.display_name)
                await checkbutton_callback(inputres)
                
            else:
                print([res.data])
           except:
                pass
         
@bot.command()
async def test(ctx):
    print("ss")
    print(await ctx.guild.fetch_roles())
    print(openxl[ctx.guild.name].cell(row= 1, column=1).number_format)
    
@bot.application_command(name="setup",description="setup attendent")
async def setup(ctx):
    servername = sheetname(ctx.guild.name)
    if ctx.user.guild_permissions.administrator:
        _row=checkRow(servername)
        for row in range(1, _row):
            openxl[servername].cell(row=row, column=3,value='출석안함')
            
        sheetsort(servername)#시트소트
        await ctx.response.send_message("dailyset 출석안함 done")
        
        es = discord.Embed(title="출석체크!",description="순위", colour=discord.Colour.blue())
        es.set_footer(text="출석하려면 버튼을 눌러주세요")
        
        _row = checkRow(servername)
        for row in range(1, _row):
                 value2 = "누적 {} 회 오늘은 {}\n마지막 출석 {}".format (openxl[servername].cell(row, 2).value, openxl[servername].cell(row, 3).value, openxl[servername].cell(row, 4).value)
                 name = get_username(str(openxl[servername].cell(row, 1).value), ctx.guild)
                 es.add_field(name=name, value=value2, inline=False)
        
        includech = ctx.guild.text_channels
        for ch in includech:
            if ch.name == "네모봇출석체크방-beta":
                        target_message=await ch.history(limit=1).flatten()
                        await target_message[0].edit(embed=es) 
                        break
        else:
            await ctx.send("전용채널이 없습니다")
    
    else:
         await ctx.response.send_message("관리자 권한이 필요합니다",ephemeral=True)

        
@bot.application_command(name="전용채널생성",description="전용채널을 생성합니다")
async def 전용채널생성(ctx): 
  if ctx.channel.name != "네모봇출석체크방-beta" and ctx.user.guild_permissions.administrator:
    
    role = discord.utils.get(ctx.guild.roles, name='네모봇')
    overwrites = {
            ctx.guild.default_role: discord.PermissionOverwrite(send_messages=False),
            role: discord.PermissionOverwrite(send_messages=True)
        }
    
    await ctx.response.send_message("명령어 사용!")

    
    h = False
    t = False
      
    includech = ctx.guild.text_channels
    for ch in includech:
              if ch.name == '네모봇출석체크방-beta':
                  await ctx.send(embed=discord.Embed(title="전용채널이 존재합니다",description=ch.name, color=0xff0000))
                  h = True
                  break
              
         
    includect = ctx.guild.categories
    for ct in includect:
              if ct.name == '네모봇':
                  await ctx.send(embed=discord.Embed(title="카테고리가 존재합니다",description=ct.name, color= 0xff0000))
                  t = True
                  break
        
    if h == False and t == True:
          category = discord.utils.get(ctx.guild.categories, name = '네모봇')
          await ctx.guild.create_text_channel('네모봇출석체크방-beta', category = category, overwrites=overwrites, topic ="출석체크방입니다\n(※알림 꺼두세요※)")
          await ctx.send(embed=discord.Embed(title="채널을 생성했습니다",description="네모봇출석체크방-beta", color= 0xff000))
      
    elif h == False and t == False:
          await ctx.guild.create_category('네모봇')
          
          category = discord.utils.get(ctx.guild.categories, name = '네모봇')
          await ctx.guild.create_text_channel('네모봇출석체크방-beta', category = category, overwrites=overwrites, topic ="출석체크방입니다\n(※알림 꺼두세요※)")
          
          embed = discord.Embed(title = "채널 및 카테고리를 생성했습니다", color = 0xff000)        
          embed.add_field(name = "생성된 카테고리", value = "네모봇")
          embed.add_field(name = "생성된 채널", value = "네모봇출석체크방-beta")
          await ctx.send(embed = embed)
      
    elif h ==  True and t == False:
          ch = discord.utils.get(ctx.guild.channels, name = '네모봇출석체크방-beta')
          await ch.delete()
          category = await ctx.guild.create_category("네모봇")
          await ctx.guild.create_text_channel('네모봇출석체크방-beta', category = category, overwrites=overwrites, topic ="출석체크방입니다n(※알림 꺼두세요※)")          
          await ctx.send(embed=discord.Embed(title="투표 채널을 초기화하고 카테고리를 생성했습니다",description="네모봇", color= 0xffff00))
      
    elif h == True and t == True:
           ch = discord.utils.get(ctx.guild.channels, name = '네모봇출석체크방-beta')
           await ch.delete()
           category = discord.utils.get(ctx.guild.categories, name = '네모봇')
           await ctx.guild.create_text_channel('네모봇출석체크방-beta', category = category, overwrites=overwrites, topic ="출석체크방입니다\n(※알림 꺼두세요※)")
           await ctx.send(embed=discord.Embed(title="투표채널을 초기화하였습니다",description="네모봇출석체크방-beta", color= 0xffff00))
          
       
    await ctx.send("투표가 생성되니다") 
    await ctx.send("투표가 생성된 방의 알림을 꺼주세요")
    
  
  elif ctx.user.guild_permissions.administrator is False:
        await ctx.response.send_message("관리자가 아닙니다",ephemeral=True)
        
  else:
     await ctx.response.send_message("현재 카테고리 제외 다른 채널에 입력해주세요",ephemeral=True)
     
     

@bot.application_command(name="전용채널삭제",description="전용채널을 삭제합니다")
async def 전용채널삭제(ctx):
 if ctx.channel.name != "네모봇출석체크방-beta" and ctx.user.guild_permissions.administrator:
        await ctx.response.send_message("명령어 사용!")
        d = 0
        
        includedch =  ctx.guild.text_channels
        for ch in includedch:
            if ch.name == '네모봇출석체크방-beta':
                     deletech = discord.utils.get(ctx.guild.channels, name = '네모봇출석체크방-beta')
                     await deletech.delete()
                     await ctx.send(embed=discord.Embed (title="채널삭제됨",description=ch.name, color= 0xffff00))
                     d=1
                     
                     
        includedct = ctx.guild.categories
        for ct in includedct:
             if ct.name == '네모봇':
                     deletect = discord.utils.get(ctx.guild.categories, name = '네모봇')
                     await deletect.delete()
                     await ctx.send(embed=discord.Embed (title="카테고리 삭제됨",description=ct.name, color= 0xffff00))
                     d=1
        
        if d==0:
                 await ctx.send(embed=discord.Embed (title="삭제할 채널 혹은 카테고리가 존재하지 않습니다",color= 0xff0000))
 
 elif  ctx.channel.name != "네모봇출석체크방-beta" and ctx.user.guild_permissions.administrator is False:
     ctx.response.send_message("관리자 권한이 필요합니다",ephemeral=True)
 
 else:
    await ctx.send("현재 카테고리를 제외한 다른 채널에 입력해주세요",ephemeral=True)
                                                                
                                   
@bot.event
async def on_guild_join(guild): 
    for joiningguild in bot.guilds:
        roles = await joiningguild.fetch_roles()
        for role in roles:
            if role.name == '네모봇':
                 await joiningguild.get_member(bot.user.id).add_roles(role)
                 break
        else:
             await joiningguild.create_role( name='네모봇', permissions= discord.Permissions(administrator=True), colour=discord.Colour.gold(), hoist=True, mentionable=True, reason="for nemobot attendent channel authority")
             await joiningguild.get_member(bot.user.id).add_roles(discord.utils.get(guild.roles, name='네모봇'))
    else:
         pass

    h = False
    t = False
    
    role = discord.utils.get(guild.roles, name='네모봇')
    overwrites = {
            guild.default_role: discord.PermissionOverwrite(send_messages=False),
            role: discord.PermissionOverwrite(send_messages=True)
        }
    
    
    includech = guild.text_channels
    for ch in includech:
              if ch.name == '네모봇출석체크방-beta':                 
                  h = True
                  break
              
         
    includect = guild.categories
    for ct in includect:
              if ct.name == '네모봇':
                  t = True
                  break
        
    if h == False and t == True:
          category = discord.utils.get(guild.categories, name = '네모봇')
          await guild.create_text_channel('네모봇출석체크방-beta', category = category, overwrites=overwrites, topic ="출석체크방입니다\n(※알림 꺼두세요※)")
          
      
    elif h == False and t == False:
          await guild.create_category('네모봇')
          
          category = discord.utils.get(guild.categories, name = '네모봇')
          await guild.create_text_channel('네모봇출석체크방-beta', category = category, overwrites=overwrites, topic ="출석체크방입니다\n(※알림 꺼두세요※)")
          
          
      
    elif h ==  True and t == False:
          ch = discord.utils.get(guild.channels, name = '네모봇출석체크방-beta')
          await ch.delete()
          category = await guild.create_category("네모봇")
          await guild.create_text_channel('네모봇출석체크방-beta', category = category, overwrites=overwrites, topic ="출석체크방입니다\n(※알림 꺼두세요※)")          
          
      
    elif h == True and t == True:
           ch = discord.utils.get(guild.channels, name = '네모봇출석체크방-beta')
           await ch.delete()
           category = discord.utils.get(guild.categories, name = '네모봇')
           await guild.create_text_channel('네모봇출석체크방-beta', category = category, overwrites=overwrites, topic ="출석체크방입니다\n(※알림 꺼두세요※)")


@bot.event
async def on_guild_channel_create(channel):
  if channel.name=="네모봇출석체크방-beta":
    value1,value2 = await bannersendfnc(channel)
    global bannersend
    bannersend= await channel.send(embed=value1, view= value2)
    
@bot.event
async def checkbutton_callback(interaction):
    ctx = interaction.guild.name
    servername = sheetname(ctx)
    _row=checkRow(ctx)
    for row in range(1, _row+1):
       ctx = interaction.guild.name
       if openxl[servername].cell(row, 1).value == str(interaction.user.id) and openxl[servername].cell(row=row, column=3).value == '출석안함':            
            n = openxl[servername].cell(row=row, column=2).value + 1
            openxl[servername].cell(row=row, column=2, value=n)
            openxl[servername].cell(row=row, column=3, value = '출석완료')
            openxl[servername].cell(row=row, column=4, value = "{}년 {}월 {}일".format (datetime.now().year, datetime.now().month, datetime.now().day))
            author = interaction.user
            pfp = author.avatar
            esb = discord.Embed(description="지금",  colour=discord.Colour.blue())
            esb.set_thumbnail(url=pfp)
            esb.add_field(name="{}".format(author),value="출석함!",inline=False)          
            await interaction.response.edit_message(embed = esb)
            time.sleep(1)
            esb.set_footer(text="출석하려면 버튼을 눌러주세요")
            await interaction.edit_original_response(embed=editbnr(interaction.channel))
            break
            
       elif openxl[servername].cell(row, 1).value == str(interaction.user.id) and openxl[servername].cell(row=row, column=3).value == '출석완료':
            embedmessage2 = discord.Embed(title="오늘 이미 출석했습니다", description=interaction.user.name)
            embedmessage2.set_footer(text="잠시뒤 메세지가 자동으로 삭제됩니다")
            await interaction.response.send_message(embed=embedmessage2, ephemeral=True, delete_after=3)
            break
            
    else:
       ctx = interaction.guild.name
       openxl[servername].cell(row= row, column=1, value= str(interaction.user.id))
       openxl[servername].cell(row= row, column=2, value=1)
       openxl[servername].cell(row= row, column=3, value = '출석완료')
       openxl[servername].cell(row= row, column=4, value = "{}년 {}월 {}일".format (datetime.now().year, datetime.now().month, datetime.now().day))
       if openxl[servername].cell(row= 1, column=1).value is None:
           openxl[servername].delete_rows(1)
       author = interaction.user
       pfp = author.avatar
       esb = discord.Embed(description="지금",  colour=discord.Colour.blue())
       esb.set_thumbnail(url=pfp)
       esb.add_field(name="{}".format(author),value="신규 출석함!",inline=False)
       await interaction.response.edit_message(embed=esb)
       time.sleep(1)
       esb.set_footer(text="출석하려면 버튼을 눌러주세요")
       await interaction.edit_original_response(embed=editbnr(interaction.channel))


@bot.event
async def on_member_update(before, after):
    for ch in before.guild.text_channels:
        if ch.name == "네모봇출석체크방-beta":
                      checkbutton = Button(label="출석확인", style=discord.ButtonStyle.green, custom_id= 'checkbutton')
    
                      view = View()
                      view.add_item(checkbutton)    
                      target_message=await ch.history(limit=None).flatten()
                      await ch. delete_messages(target_message)
                      global bannersend
                      bannersend=await ch.send(embed=editbnr(ch), view=view)
                      break
    else:
        pass   


bot.run("MTE3MDY5MDg5MTk3NjAyNDA5NA.G-YeAh.meyfKhDlJyplFWXVKl6mGeaO4vM6niSh4c0RU4")
